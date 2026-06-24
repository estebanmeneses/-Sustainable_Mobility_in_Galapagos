import numpy as np
import pandas as pd
from openpyxl import load_workbook

import pickup_driving_cycle as driving
import power_cycle as power
import state_of_charge as soc

ACTIVITY_FILE = "activity_profile.xlsx"
FLEET_FILE    = "Prospectiva_Transporte_T_Galapagos.xlsx"
YEARS         = list(range(2019, 2051))
BAND_SIZE     = 3 * 3600  # segundos por franja de 3h
BAND_LABELS   = [f"{h:02d}:00-{h+3:02d}:00" for h in range(0, 24, 3)]

VEHICLE_SHEETS = ["PickUp", "SUV", "Bus", "Heavy Truck", "Motorcycle", "Micromobilidad"]

FLEET_NAME_MAP = {
    "PickUp":         "Camioneta (taxi)-electrica",
    "SUV":            "Vehículo eléctrico -electricidad",
    "Bus":            "Bus y furgoneta - electrica",
    "Motorcycle":     "Motocicleta-electrica",
    "Micromobilidad": "Micromo",
}

HEAVY_TRUCK_ROWS = [
    "Camión- electrico",
    "Especial (tanqueros, volquetas, etc.)-electrica",
]

SCENARIOS = {
    "Demanda_SC_low":    "Flota_low_text",
    "Demanda_SC_medium": "Flota_medium_text",
    "Demanda_SC_high":   "Flota_high_text",
}


def compute_per_vehicle_band_kwh() -> dict:
    """Corre el pipeline una vez por tipo y retorna kWh por vehículo para cada franja de 3h."""
    band_kwh = {}
    for sheet in VEHICLE_SHEETS:
        print(f"  Simulando: {sheet}")
        v, frr, theta = driving.generate_driving_profile(ACTIVITY_FILE, sheet)
        p_req, battery_cap = power.calculate_power(v, frr, theta, ACTIVITY_FILE, sheet)
        _, _, _, _, _, grid_p = soc.calculate_soc_with_charging(p_req, battery_cap, ACTIVITY_FILE, sheet)
        band_kwh[sheet] = [
            float(np.sum(grid_p[i * BAND_SIZE:(i + 1) * BAND_SIZE]) / 3600)
            for i in range(8)
        ]
        print(f"    kWh/vehículo/día: {sum(band_kwh[sheet]):.3f}")
    return band_kwh


def read_santa_cruz_section(fleet_sheet_name: str) -> pd.DataFrame:
    """Devuelve la sección de Santa Cruz como DataFrame (índice = nombre vehículo, cols = años)."""
    df = pd.read_excel(FLEET_FILE, sheet_name=fleet_sheet_name)
    col0 = df.columns[0]

    sc_mask = df[col0].astype(str).str.contains("Santa Cruz", na=False)
    sc_pos  = int(df[sc_mask].index[0])

    section = df.iloc[sc_pos + 1: sc_pos + 17].copy()
    section = section.set_index(col0)

    year_cols = [c for c in section.columns if isinstance(c, (int, float)) and 2019 <= int(c) <= 2050]
    section   = section[year_cols].fillna(0)
    section.columns = [int(c) for c in section.columns]

    return section


def get_counts(section: pd.DataFrame, fleet_name: str) -> np.ndarray:
    """Retorna array de counts (shape 32) para el nombre de vehículo dado."""
    idx = section.index.astype(str)
    matches = [i for i in idx if fleet_name.strip().lower() in i.strip().lower()]
    if not matches:
        print(f"    ADVERTENCIA: '{fleet_name}' no encontrado en datos de Santa Cruz")
        return np.zeros(len(YEARS))
    return section.loc[matches[0]].reindex(YEARS, fill_value=0).values.astype(float)


def build_demand_matrix(band_kwh: dict, fleet_sheet_name: str) -> np.ndarray:
    """Construye la matriz de demanda (8 franjas × 32 años) en kWh para Santa Cruz."""
    section = read_santa_cruz_section(fleet_sheet_name)
    demand  = np.zeros((8, len(YEARS)))

    for sheet in VEHICLE_SHEETS:
        if sheet == "Heavy Truck":
            counts = (get_counts(section, HEAVY_TRUCK_ROWS[0]) +
                      get_counts(section, HEAVY_TRUCK_ROWS[1]))
        else:
            counts = get_counts(section, FLEET_NAME_MAP[sheet])

        for i in range(8):
            demand[i] += band_kwh[sheet][i] * counts

    return demand


def write_to_excel(demand: np.ndarray, output_sheet_name: str) -> None:
    """Escribe la matriz de demanda como nueva hoja en el Excel de flota."""
    wb = load_workbook(FLEET_FILE)

    if output_sheet_name in wb.sheetnames:
        del wb[output_sheet_name]

    ws = wb.create_sheet(output_sheet_name)

    # Fila 1: encabezado
    ws.append(["Demanda SC"] + YEARS)

    # Fila 2: "Santa Cruz" + energía diaria total por año (suma de las 8 franjas)
    daily_totals = demand.sum(axis=0).tolist()
    ws.append(["Santa Cruz"] + [round(v, 4) for v in daily_totals])

    # Filas 3-10: energía por franja horaria
    for i, label in enumerate(BAND_LABELS):
        ws.append([label] + [round(v, 4) for v in demand[i]])

    wb.save(FLEET_FILE)
    print(f"  OK Hoja '{output_sheet_name}' guardada.")


def plot_per_scenario_demand(all_demands: dict) -> None:
    """Una gráfica por escenario: demanda del día 2026 en 8 franjas horarias."""
    import matplotlib.pyplot as plt
    import matplotlib.cm as cm

    year_idx = YEARS.index(2026)
    band_colors = [cm.RdYlGn(i / 7) for i in range(8)]
    x = np.arange(8)
    titles = {
        "Demanda_SC_low":    "Escenario Low — Demanda Energetica Santa Cruz 2026",
        "Demanda_SC_medium": "Escenario Medium — Demanda Energetica Santa Cruz 2026",
        "Demanda_SC_high":   "Escenario High — Demanda Energetica Santa Cruz 2026",
    }

    for name, demand in all_demands.items():
        values = demand[:, year_idx]
        _, ax = plt.subplots(figsize=(11, 6))

        ax.bar(x, values, width=1.0, color=band_colors, edgecolor="none")

        for i, val in enumerate(values):
            if val > 0:
                ax.text(i, val + values.max() * 0.01, f"{val:.1f}",
                        ha="center", va="bottom", fontsize=9, fontweight="bold")

        total = values.sum()
        ax.text(0.98, 0.97, f"Total diario: {total:.1f} kWh",
                transform=ax.transAxes, ha="right", va="top", fontsize=10,
                bbox=dict(facecolor="white", alpha=0.85, edgecolor="gray"))

        ax.set_title(titles.get(name, name), fontsize=13)
        ax.set_xlabel("Franja Horaria", fontsize=11)
        ax.set_ylabel("Energia Demandada (kWh)", fontsize=11)
        ax.set_xticks(x)
        ax.set_xticklabels(BAND_LABELS, rotation=20, ha="right", fontsize=10)
        ax.set_ylim(0, values.max() * 1.18)
        ax.grid(True, linestyle="--", alpha=0.4, axis="y")
        plt.tight_layout()
        plt.show()


def plot_2026_demand(all_demands: dict) -> None:
    """Gráfico de barras agrupadas: demanda energética de la flota en 2026 por escenario."""
    import matplotlib.pyplot as plt

    year_idx = YEARS.index(2026)
    x = np.arange(8)
    width = 0.25
    colors = {
        "Demanda_SC_low":    "#5dade2",
        "Demanda_SC_medium": "#f39c12",
        "Demanda_SC_high":   "#27ae60",
    }
    labels = {
        "Demanda_SC_low":    "Escenario Low",
        "Demanda_SC_medium": "Escenario Medium",
        "Demanda_SC_high":   "Escenario High",
    }
    offsets = [-1, 0, 1]

    _, ax = plt.subplots(figsize=(13, 7))

    for (name, demand), offset in zip(all_demands.items(), offsets):
        values = demand[:, year_idx]
        ax.bar(x + offset * width, values, width=width,
               label=labels[name], color=colors[name],
               alpha=0.88, edgecolor="gray", linewidth=0.5)

    ax.set_title("Demanda Energetica Flota Santa Cruz - 2026 (3 Escenarios)", fontsize=14)
    ax.set_xlabel("Franja Horaria", fontsize=12)
    ax.set_ylabel("Energia Demandada (kWh)", fontsize=12)
    ax.set_xticks(x)
    ax.set_xticklabels(BAND_LABELS, rotation=20, ha="right", fontsize=10)
    ax.legend(loc="upper right", frameon=True)
    ax.grid(True, linestyle="--", alpha=0.4, axis="y")
    plt.tight_layout()
    plt.show()


if __name__ == "__main__":
    print("=== Paso 1: Calculando perfil de energia por vehiculo ===")
    band_kwh = compute_per_vehicle_band_kwh()

    all_demands = {}
    for output_name, fleet_sheet in SCENARIOS.items():
        print(f"\n=== Paso 2-4: Escenario '{output_name}' (fuente: {fleet_sheet}) ===")
        demand = build_demand_matrix(band_kwh, fleet_sheet)
        all_demands[output_name] = demand
        print(f"  Energia total 2050: {demand[:, -1].sum():.1f} kWh/dia")
        write_to_excel(demand, output_name)

    print(f"\nCompletado. Abre: {FLEET_FILE}")
    plot_per_scenario_demand(all_demands)
    plot_2026_demand(all_demands)
