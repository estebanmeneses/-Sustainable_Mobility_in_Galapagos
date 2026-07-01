from openpyxl import load_workbook, Workbook  # noqa: F401

import datetime
import numpy as np
import pandas as pd
import vehicle_cycle
import fleet_power
from config import FRR_ASPHALTED, FRR_GRAVEL
from utils import extract_island_data, to_sec

YEARS               = list(range(2019, 2051))
BAND_LABELS         = [f"{h:02d}:00-{h+3:02d}:00" for h in range(0, 24, 3)]
ISLANDS             = vehicle_cycle.ISLANDS
VEHICLE_SHEETS      = vehicle_cycle.VEHICLE_SHEETS
OUTPUT_FILE         = "output/energia_por_vehiculo.xlsx"
SOC_FILE            = "output/soc_por_vehiculo.xlsx"

def write_energy_excel(band_kwh: dict) -> None:
    import os
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["Energia"] if "Energia" in wb.sheetnames else wb.create_sheet("Energia", 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Energia"
    row_idx = 1

    def write_row(values):
        nonlocal row_idx
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1

    write_row(["Energia (kWh)"] + YEARS)

    for island in ISLANDS:
        write_row([island])
        for sheet in VEHICLE_SHEETS:
            write_row([sheet])
            for b, label in enumerate(BAND_LABELS):
                val = band_kwh[island][sheet][b]
                write_row([label] + [round(val, 4)] * len(YEARS))

    wb.save(OUTPUT_FILE)
    print(f"Excel guardado: {OUTPUT_FILE}")


def write_soc_excel(soc_ch_profiles: dict) -> None:
    import os
    os.makedirs("output", exist_ok=True)
    time_h = np.arange(86400) / 3600.0
    with pd.ExcelWriter(SOC_FILE, engine="openpyxl") as writer:
        for sheet in VEHICLE_SHEETS:
            data = {"Tiempo (h)": time_h}
            for island in ISLANDS:
                data[f"SoC {island} (%)"] = soc_ch_profiles[island][sheet] * 100
            pd.DataFrame(data).to_excel(writer, sheet_name=sheet, index=False)
    print(f"Excel SoC guardado: {SOC_FILE}")


if __name__ == "__main__":
    band_kwh, distances, grid_p_profiles, soc_ch_profiles, driving_profiles = \
        vehicle_cycle.run_multi_island_analysis()
    write_energy_excel(band_kwh)
    write_soc_excel(soc_ch_profiles)
    all_power, all_energy = fleet_power.run(grid_p_profiles)
    fleet_power.save_results(all_power, all_energy)
